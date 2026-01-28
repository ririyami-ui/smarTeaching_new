import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ATP"

# Headers
headers = ["No", "Elemen", "Lingkup Materi", "Tujuan Pembelajaran", "JP", "Profil Lulusan"]
ws.append(headers)

# Style headers
header_fill = PatternFill(start_color="7E22CE", end_color="7E22CE", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Sample data
sample_data = [
    [1, "Bilangan", "Bilangan Bulat", "Melalui diskusi kelompok, peserta didik dapat menganalisis operasi bilangan bulat dengan tepat.", 6, "Penalaran Kritis"],
    [2, "Aljabar", "Persamaan Linear", "Peserta didik dapat menyelesaikan persamaan linear satu variabel melalui latihan soal secara mandiri.", 8, "Kemandirian, Penalaran Kritis"],
    [3, "Geometri", "Bangun Datar", "Peserta didik mampu menghitung luas dan keliling bangun datar melalui praktik pengukuran.", 6, "Kreativitas, Kolaborasi"]
]

for row_data in sample_data:
    ws.append(row_data)

# Set column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 60
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 30

# Apply borders to all cells
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = thin_border
        if cell.row > 1:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

# Save
wb.save('F:/app-firebase/Smart Teaching/smart-teaching-manager/template/template_ATP.xlsx')
print("Template ATP berhasil dibuat!")
